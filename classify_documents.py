"""
classify_documents.py — Complexity classifier for district compensation plan documents

Reads Best_URL from the Unique Districts tab, fetches each document, and classifies it:
  Simple       — PDF/XLSX ≤3 pages, salary keywords found, at least one table
  Medium       — PDF/XLSX 4-15 pages, salary keywords found
  Complex      — PDF/XLSX >15 pages OR dense multi-table structure
  HTML         — Resolved to an HTML page (not a downloadable doc)
  Wrong Doc    — Fetched OK but no salary-related keywords found
  Unreadable   — PDF present but no extractable text (likely scanned image)
  Skipped      — QA_Status not a success indicator (dead/timeout/etc.)
  Error        — Fetch or parse exception

Writes 4 new columns to Unique Districts tab starting at col P (16):
  P: Doc_Class   — classification label
  Q: Doc_Pages   — page count (PDF/XLSX only)
  R: Doc_Tables  — table count detected
  S: Doc_Notes   — detail (keyword sample, error, etc.)

Resume: skips rows where Doc_Class (col P) is already populated.
Redirect: if Redirect_URL (col O) is non-empty and QA_Status contains REDIRECT, uses that URL.

Usage:
    python classify_documents.py [--dry-run] [--limit N] [--start-row N] [--end-row N]
"""

import argparse
import io
import logging
import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    print("WARNING: pdfplumber not installed. PDF classification disabled. Run: pip install pdfplumber")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

load_dotenv()

# ── Constants ─────────────────────────────────────────────────────────────────

SPREADSHEET_ID  = os.getenv("SPREADSHEET_ID")
CREDS_JSON      = os.getenv("GOOGLE_SHEETS_CREDS_JSON")
SHEET_NAME      = "Unique Districts"

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Column indices (1-based)
COL_DIST_NAME    = 4   # D
COL_BEST_URL     = 10  # J
COL_QA_STATUS    = 14  # N
COL_REDIRECT_URL = 15  # O
COL_DOC_CLASS    = 16  # P  ← new
COL_DOC_PAGES    = 17  # Q  ← new
COL_DOC_TABLES   = 18  # R  ← new
COL_DOC_NOTES    = 19  # S  ← new

WRITE_COLS = {COL_DOC_CLASS, COL_DOC_PAGES, COL_DOC_TABLES, COL_DOC_NOTES}
WRITE_COL_MIN = COL_DOC_CLASS   # 16  P
WRITE_COL_MAX = COL_DOC_NOTES   # 19  S

NEW_HEADERS = ["Doc_Class", "Doc_Pages", "Doc_Tables", "Doc_Notes"]

FETCH_TIMEOUT   = 20   # seconds
FETCH_MAX_BYTES = 30 * 1024 * 1024  # 30 MB cap — skip absurdly large files
SLEEP_BETWEEN   = 1.2  # seconds between fetches
BATCH_SIZE      = 20

QA_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0.0.0 Safari/537.36"
)

# Keywords that indicate a salary/compensation document
SALARY_KEYWORDS = [
    r"salary",
    r"compensation",
    r"pay\s+scale",
    r"pay\s+schedule",
    r"pay\s+plan",
    r"\bwage\b",
    r"\bstipend\b",
    r"\bstep\b",
    r"\blane\b",
    r"base\s+pay",
    r"hourly\s+rate",
    r"annual\s+pay",
    r"teacher\s+pay",
    r"salary\s+schedule",
]
SALARY_RE = re.compile("|".join(SALARY_KEYWORDS), re.IGNORECASE)

# QA status prefixes that indicate a fetchable document
FETCHABLE_QA_PREFIXES = ("✓ PDF", "✔ PDF", "✓ XLSX", "✔ XLSX", "⚠ REDIRECT")


# ── Logging ───────────────────────────────────────────────────────────────────

def setup_logging() -> logging.Logger:
    Path("logs").mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = f"logs/classify_{ts}.log"

    logger = logging.getLogger("classifier")
    logger.setLevel(logging.DEBUG)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))

    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info(f"Log: {log_path}")
    return logger


# ── Google Sheets ──────────────────────────────────────────────────────────────

def open_sheet(logger):
    if not GSPREAD_AVAILABLE:
        logger.error("gspread not installed.")
        sys.exit(1)
    creds = Credentials.from_service_account_file(CREDS_JSON, scopes=SCOPES)
    gc = gspread.authorize(creds)
    wb = gc.open_by_key(SPREADSHEET_ID)
    ws = wb.worksheet(SHEET_NAME)
    return ws


def ensure_headers(ws, dry_run: bool, logger):
    """Write P1:S1 headers if blank."""
    current = ws.row_values(1)
    # Extend list if shorter than col S (19)
    while len(current) < COL_DOC_NOTES:
        current.append("")
    needs_write = any(
        current[COL_DOC_CLASS - 1 + i] == ""
        for i in range(4)
    )
    if needs_write:
        if dry_run:
            logger.info(f"[dry-run] Would write headers {NEW_HEADERS} to P1:S1")
        else:
            ws.update(
                values=[NEW_HEADERS],
                range_name="P1:S1",
                value_input_option="USER_ENTERED",
            )
            logger.info("Wrote headers P1:S1")


def flush_batch(ws, buffer: list, dry_run: bool, logger):
    if not buffer:
        return
    if dry_run:
        for item in buffer:
            logger.info(
                f"  [dry-run] Row {item['row']}: "
                f"Class={item['cls']} Pages={item['pages']} "
                f"Tables={item['tables']} Notes={item['notes'][:60]}"
            )
        buffer.clear()
        return

    data = []
    for item in buffer:
        r = item["row"]
        data.append({
            "range": f"P{r}:S{r}",
            "values": [[item["cls"], item["pages"], item["tables"], item["notes"]]],
        })
    ws.batch_update(data, value_input_option="USER_ENTERED")
    buffer.clear()


# ── Fetch ──────────────────────────────────────────────────────────────────────

def fetch_bytes(url: str) -> tuple[bytes | None, str | None]:
    """
    Returns (content_bytes, content_type) or (None, error_msg).
    Streams response and caps at FETCH_MAX_BYTES.
    """
    headers = {"User-Agent": QA_USER_AGENT}
    try:
        resp = requests.get(
            url, headers=headers, timeout=FETCH_TIMEOUT,
            stream=True, allow_redirects=True,
        )
        ct = resp.headers.get("Content-Type", "").lower()

        if resp.status_code >= 400:
            return None, f"HTTP {resp.status_code}"

        chunks = []
        total = 0
        for chunk in resp.iter_content(chunk_size=65536):
            chunks.append(chunk)
            total += len(chunk)
            if total > FETCH_MAX_BYTES:
                return None, f"File too large (>{FETCH_MAX_BYTES // (1024*1024)} MB)"

        return b"".join(chunks), ct
    except requests.exceptions.Timeout:
        return None, "Timeout"
    except Exception as e:
        return None, str(e)[:120]


# ── Classification helpers ────────────────────────────────────────────────────

def has_salary_keywords(text: str) -> bool:
    return bool(SALARY_RE.search(text))


def keyword_sample(text: str, n: int = 3) -> str:
    """Return up to n unique matched keywords for notes."""
    found = set()
    for pat in SALARY_KEYWORDS:
        if re.search(pat, text, re.IGNORECASE):
            # Use the pattern stem as the label
            found.add(pat.replace(r"\b", "").replace(r"\s+", " ").strip())
        if len(found) >= n:
            break
    return ", ".join(sorted(found)) if found else ""


def classify_pdf(content: bytes) -> dict:
    """
    Returns dict with keys: cls, pages, tables, notes
    """
    if not PDFPLUMBER_AVAILABLE:
        return {"cls": "Unreadable", "pages": "", "tables": "", "notes": "pdfplumber not installed"}

    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            page_count = len(pdf.pages)
            all_text = ""
            table_count = 0

            for page in pdf.pages:
                # Extract text
                page_text = page.extract_text() or ""
                all_text += page_text + "\n"

                # Count tables on this page
                tables = page.extract_tables()
                table_count += len(tables) if tables else 0

        if not all_text.strip():
            return {
                "cls": "Unreadable",
                "pages": page_count,
                "tables": table_count,
                "notes": "No extractable text (likely scanned image)",
            }

        if not has_salary_keywords(all_text):
            return {
                "cls": "Wrong Doc",
                "pages": page_count,
                "tables": table_count,
                "notes": f"No salary keywords found",
            }

        # Classify by page count
        if page_count <= 3:
            cls = "Simple"
        elif page_count <= 15:
            cls = "Medium"
        else:
            cls = "Complex"

        sample = keyword_sample(all_text)
        notes = f"keywords: {sample}" if sample else ""

        return {"cls": cls, "pages": page_count, "tables": table_count, "notes": notes}

    except Exception as e:
        return {"cls": "Error", "pages": "", "tables": "", "notes": str(e)[:120]}


def classify_html(content: bytes) -> dict:
    """Classify an HTML page."""
    try:
        soup = BeautifulSoup(content, "html.parser")
        text = soup.get_text(separator=" ")

        table_count = len(soup.find_all("table"))

        if not has_salary_keywords(text):
            return {
                "cls": "Wrong Doc",
                "pages": "",
                "tables": table_count,
                "notes": "No salary keywords found in HTML page",
            }

        sample = keyword_sample(text)
        notes = f"HTML page; keywords: {sample}" if sample else "HTML page"
        return {"cls": "HTML", "pages": "", "tables": table_count, "notes": notes}

    except Exception as e:
        return {"cls": "Error", "pages": "", "tables": "", "notes": str(e)[:120]}


def classify_xlsx(content: bytes) -> dict:
    """Classify an Excel file."""
    if not OPENPYXL_AVAILABLE:
        return {"cls": "Error", "pages": "", "tables": "", "notes": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet_count = len(wb.sheetnames)
        all_text = ""
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        all_text += str(cell) + " "

        if not has_salary_keywords(all_text):
            return {
                "cls": "Wrong Doc",
                "pages": sheet_count,
                "tables": "",
                "notes": "No salary keywords found in XLSX",
            }

        sample = keyword_sample(all_text)
        notes = f"XLSX; keywords: {sample}" if sample else "XLSX"
        # Use sheet count as rough complexity proxy
        cls = "Simple" if sheet_count <= 2 else "Medium"
        return {"cls": cls, "pages": sheet_count, "tables": "", "notes": notes}

    except Exception as e:
        return {"cls": "Error", "pages": "", "tables": "", "notes": str(e)[:120]}


def classify_url(url: str, qa_status: str, redirect_url: str) -> dict:
    """
    Resolve the final URL to use, fetch it, and classify.
    """
    # Use redirect target if available
    target_url = url
    if redirect_url and redirect_url.strip():
        target_url = redirect_url.strip()

    content, ct_or_err = fetch_bytes(target_url)
    if content is None:
        return {"cls": "Error", "pages": "", "tables": "", "notes": ct_or_err}

    ct = ct_or_err or ""

    # Detect type from Content-Type or URL extension
    url_lower = target_url.lower().split("?")[0]

    is_pdf = (
        "pdf" in ct
        or url_lower.endswith(".pdf")
    )
    is_xlsx = (
        any(x in ct for x in ["spreadsheet", "excel", "openxml"])
        or url_lower.endswith(".xlsx")
        or url_lower.endswith(".xls")
    )
    is_html = "html" in ct or url_lower.endswith(".html") or url_lower.endswith(".htm")

    if is_pdf:
        return classify_pdf(content)
    elif is_xlsx:
        return classify_xlsx(content)
    elif is_html:
        return classify_html(content)
    else:
        # Unknown type — attempt PDF then HTML
        result = classify_pdf(content)
        if result["cls"] == "Error":
            result = classify_html(content)
        result["notes"] = f"[ambiguous ct={ct[:40]}] " + result.get("notes", "")
        return result


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Classify district compensation plan documents.")
    parser.add_argument("--dry-run",      action="store_true", help="Print what would be written; no sheet writes")
    parser.add_argument("--limit",        type=int, default=0, help="Process first N eligible rows only")
    parser.add_argument("--random",       type=int, default=0, metavar="N", help="Randomly sample N eligible rows (ignores --limit/--start-row/--end-row)")
    parser.add_argument("--start-row",    type=int, default=2, help="Sheet row to start from (default 2)")
    parser.add_argument("--end-row",      type=int, default=0, help="Sheet row to stop at (0 = all rows)")
    parser.add_argument("--rerun-errors", action="store_true", help="Re-classify only rows where Doc_Class (col P) is 'Error', bypassing resume")
    args = parser.parse_args()

    logger = setup_logging()

    if args.dry_run:
        logger.info("=== DRY RUN — no sheet writes, no fetches ===")

    # ── Auth ──────────────────────────────────────────────────────────────────
    if args.dry_run:
        ws = None
        logger.info("[dry-run] Skipping Sheets auth")
    else:
        ws = open_sheet(logger)
        ensure_headers(ws, dry_run=False, logger=logger)

    # ── Load rows ─────────────────────────────────────────────────────────────
    if args.dry_run:
        # Mock 5 rows for demo
        mock_rows = [
            # (row_num, dist_name, best_url, qa_status, redirect_url, existing_class)
            (2,  "Austin ISD",      "https://www.austinisd.org/sites/default/files/dept/hr/25-26-comp-plan.pdf", "✓ PDF",  "", ""),
            (3,  "Wimberley ISD",   "https://www.wimberleyisd.net/improvement-plan-2025.pdf", "✓ PDF", "", ""),
            (4,  "Houston ISD",     "https://www.houstonisd.org/compensation2025.pdf", "✓ PDF", "", ""),
            (5,  "Midland ISD",     "https://www.midlandisd.net/compensation", "✓ HTML", "", ""),
            (6,  "Schertz-Cibolo",  "https://example.com/dead.pdf", "✗ DEAD", "", ""),
        ]
        all_values = mock_rows
    else:
        raw = ws.get_all_values()
        # raw[0] is header row (row 1), raw[1] is row 2, etc.
        all_values = []
        for i, row in enumerate(raw[1:], start=2):  # start=2 because row 1 is headers
            def cell(col): return row[col - 1] if len(row) >= col else ""
            all_values.append((
                i,
                cell(COL_DIST_NAME),
                cell(COL_BEST_URL),
                cell(COL_QA_STATUS),
                cell(COL_REDIRECT_URL),
                cell(COL_DOC_CLASS),   # existing classification (for resume)
            ))

    # Apply row range filters
    start = args.start_row
    end   = args.end_row or 999999
    all_values = [(r, *rest) for r, *rest in all_values if start <= r <= end]

    # Random sampling: pick N rows from those with a fetchable URL (pre-filter)
    if args.random:
        eligible = [
            row for row in all_values
            if not row[5].strip()                              # not already classified
            and row[2].strip()                                 # has a URL
            and any(row[3].startswith(p) for p in FETCHABLE_QA_PREFIXES)  # QA ok
        ]
        sample_n = min(args.random, len(eligible))
        sampled = random.sample(eligible, sample_n)
        sampled_rows = {row[0] for row in sampled}
        # Keep sampled rows plus non-fetchable rows (they get written as Skipped)
        all_values = [row for row in all_values if row[0] in sampled_rows or not row[2].strip() or not any(row[3].startswith(p) for p in FETCHABLE_QA_PREFIXES)]
        # But we only want the sampled fetchable ones — drop non-fetchable entirely for a clean run
        all_values = sampled
        logger.info(f"Random sample: {sample_n} rows selected from {len(eligible)} eligible")

    # ── Process ───────────────────────────────────────────────────────────────
    buffer = []
    processed = 0
    skipped_resume = 0
    skipped_qa = 0

    for row_num, dist_name, best_url, qa_status, redirect_url, existing_class in all_values:
        # Resume: skip already classified rows, unless --rerun-errors targets Error rows
        if existing_class.strip():
            if args.rerun_errors and existing_class.strip() == "Error":
                pass  # fall through and re-classify
            else:
                skipped_resume += 1
                continue

        # Skip rows with no URL or failed QA
        if not best_url or not best_url.strip():
            skipped_qa += 1
            continue

        qa_ok = any(qa_status.startswith(p) for p in FETCHABLE_QA_PREFIXES)
        if not qa_ok:
            result = {"cls": "Skipped", "pages": "", "tables": "", "notes": f"QA: {qa_status[:40]}"}
            buffer.append({
                "row": row_num, "cls": result["cls"],
                "pages": result["pages"], "tables": result["tables"], "notes": result["notes"],
            })
            skipped_qa += 1
            # Flush these passively with the normal batch
        else:
            if args.dry_run:
                # Don't actually fetch in dry-run — mock a result based on district name
                if "wimberley" in dist_name.lower():
                    result = {"cls": "Wrong Doc", "pages": 8, "tables": 0,
                              "notes": "No salary keywords found"}
                elif "houston" in dist_name.lower():
                    result = {"cls": "Complex", "pages": 42, "tables": 18,
                              "notes": "keywords: salary, compensation, step"}
                elif "austin" in dist_name.lower():
                    result = {"cls": "Simple", "pages": 2, "tables": 3,
                              "notes": "keywords: salary, step, lane"}
                elif "midland" in dist_name.lower():
                    result = {"cls": "HTML", "pages": "", "tables": 4,
                              "notes": "HTML page; keywords: salary, pay scale"}
                else:
                    result = {"cls": "Skipped", "pages": "", "tables": "",
                              "notes": f"QA: {qa_status}"}

                logger.info(
                    f"  [dry-run] Row {row_num} | {dist_name[:35]:<35} | "
                    f"{result['cls']:<12} | pages={result['pages']} tables={result['tables']}"
                )
            else:
                logger.info(f"[Row {row_num}] {dist_name[:40]} — fetching {best_url[:60]}...")
                result = classify_url(best_url, qa_status, redirect_url)
                logger.info(
                    f"  -> {result['cls']} | pages={result['pages']} "
                    f"tables={result['tables']} | {result['notes'][:60]}"
                )
                time.sleep(SLEEP_BETWEEN)

            buffer.append({
                "row": row_num, "cls": result["cls"],
                "pages": result["pages"], "tables": result["tables"], "notes": result["notes"],
            })
            processed += 1

        if len(buffer) >= BATCH_SIZE:
            flush_batch(ws, buffer, args.dry_run, logger)

        if args.limit and processed >= args.limit:
            logger.info(f"Reached --limit {args.limit}, stopping.")
            break

    # Final flush
    flush_batch(ws, buffer, args.dry_run, logger)

    logger.info(
        f"\nDone. Classified: {processed} | "
        f"Skipped (resume): {skipped_resume} | Skipped (QA/no URL): {skipped_qa}"
    )


if __name__ == "__main__":
    main()
